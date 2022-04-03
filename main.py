import math
import random
from copy import deepcopy
from openpyxl import load_workbook
from openpyxl import Workbook

wb = load_workbook('./Doswiadczenia-v03.xlsx')
sigma_test = []
e = []
T = []
e_dot = []
q_def = 312500
R_gaz = 8.314


for ws in wb.worksheets:
    e_values = [ws.cell(i, 1).value for i in range(4, 55)]
    sigma_values = [ws.cell(i, 2).value for i in range(4, 55)]
    T.append(ws.cell(2, 1))
    e_dot.append(ws.cell(2, 2))
    e.append(e_values)
    sigma_test.append(sigma_values)

smallest = 10000000000000.0
constraints = [[1.0, 0.0, 0.0, 1.0, 0.0, 1.0, 0.0], [1000.0, 1.0, 1.0, 10000.0, 1.0, 90000.0, 1.0]]
a = [0, 0, 0, 0, 0, 0, 0]
s = 0.02
# 0 < alpha < 1
alpha = 0.5
# precision
epsilon = 0.000001
result_a = []
smallest_a = []
result_objective = 0
objective_values = []
smallest_objective_values = []
for i in range(20):
    print("step " + str(i))
    for i in range(7):
        a[i] = random.uniform(constraints[0][i], constraints[1][i])
        # a = [500, 0.5, 0.5, 5000, 0.5, 45000, 0.5]
        # relative step size
    objective_values = []
    result_a = hooke_jeeves(a, s, alpha, epsilon, e, T, e_dot, sigma_test, constraints)
    result_objective = objective(result_a, e, T, e_dot, sigma_test)
    if result_objective < smallest:
        smallest_objective_values = deepcopy(objective_values)
        smallest = result_objective
        smallest_a = deepcopy(result_a)
        print(result_a)
        print(result_objective)
# print(result_a)

# ws1 = wb.create_sheet(title="Results")
wb1 = load_workbook('./Results.xlsx', read_only=False)
ws1 = wb1.worksheets[0]
ws1.cell(1, 3, result_objective)
for i in range(len(smallest_a)):
    ws1.cell(i + 1, 1, smallest_a[i])
for i in range(len(smallest_objective_values)):
    ws1.cell(i + 1, 2, smallest_objective_values[i])

for ws in wb.worksheets:
    for i in range(len(e[0])):
        e1 = e[0][i]
        T1 = ws.cell(2, 4).value
        edot1 = ws.cell(2, 5).value
        ws.cell(i + 2, 3, sigma_p(smallest_a, e1, T1, edot1))

wb1.save('./Results.xlsx')
wb.save('./Doswiadczenia.xlsx')
wb.close()
wb1.close()

# 127.41171114187962
# [179.55596276914937, 0.5, 0.12200341075755708, -1014.2046975999997, 5.8713463056566155, 34944.01339230601, 0.23720028880244573]
# 127.41216971968127
# [179.56598455445413, 0.5, 0.12206022764828782, -1014.2046975999997, 5.8713463056566155, 34938.16057412525, 0.2373760505942261]
# [169.60147108378362, 0.20798435025038958, 0.12199999663570239, 4517.219461080158, 0.3179692643007357, 11432.521198175353, 0.5337270857345702]
# 9.532441364190856e-06


def sigma_p(a, e, T, e_dot):
    R = 8.314
    W = math.exp(-a[6] * e)
    T1 = math.exp(a[3] / (R * (T + 273)))
    T2 = math.exp(a[5] / (R * (T + 273)))
    e1 = e ** a[1]
    edot1 = e_dot ** a[2]
    result = W * a[0] * T1 * e1
    result = result + ((1 - W) * a[4] * T2)
    result = result * edot1
    return result


def sigma_p2():
    Z = e_dot * math.exp(q_def * R_gaz * T_def)
    sigma0 = 1 / a[2] * (1 / math.sinh(Z / a[0]) ** (1 / a[1]))
    sigma_sse = 1 / a[5] * (1 / math.sinh(Z / a[3]) ** (1 / a[4]))
    sigma_ss = 1 / a[10] * (1 / math.sinh(Z / a[8]) ** (1 / a[9]))
    e_r = (a[6] + a[7] * sigma_sse ** 2) / 3.23
    e_xsc = a[13] * (Z / sigma_sse ** 2) ** a[14]
    e_xrc = e_xsc / 1.98
    e_c = a[11] * (Z / sigma_sse ** 2) ** a[12]
    R = 0
    if e > e_c:
        R = (sigma_sse - sigma_ss) * (1 - math.exp(-((e - e_c) / e_xrc) ** 2))
    sigma_p = sigma0 + (sigma_sse - sigma0) * math.sqrt(1 - math.exp(-e / e_r)) - R




def objective(a, e, T, e_dot, sigma_test):
    sum = 0
    for i in range(len(T)):
        for j in range(len(e[0])):
            err_squared = (sigma_test[i][j] - sigma_p(a, e[i][j], T[i], e_dot[i])) ** 2
            err_relative = err_squared / sigma_test[i][j]
            sum = sum + err_relative
    return sum

# def objective(a, data, num_of_experiments, num_of_measurements):



def hooke_jeeves(x, s, alpha, epsilon, e, T, e_dot, sigma_test, constraints):
    while s > epsilon:
        xb = deepcopy(x)
        objective_values.append(objective(x, e, T, e_dot, sigma_test))
        x = trial(xb, s, constraints)
        if objective(x, e, T, e_dot, sigma_test) < objective(xb, e, T, e_dot, sigma_test):
            while True:
                xb1 = deepcopy(xb)
                xb = deepcopy(x)
                flag = True
                temp = [2 * x for x in xb]
                for i in range(len(temp)):
                    temp[i] = temp[i] - xb1[i]
                    if not constraints[1][i] > temp[i] > constraints[0][i]:
                        flag = False
                        break
                if flag:
                    x = deepcopy(temp)
                # x = [2 * x for x in xb]
                # temp = []
                # for i in range(len(x)):
                #     temp.append(x[i] - xb1[i])
                #     # x[i] = temp
                #     if not constraints[1][i] > temp[i] > constraints[0][i]:
                #         # print('limip ipsum')
                #         break
                # else:
                #     x = deepcopy(temp)
                # x = 2 * xb - xb1
                x = trial(x, s, constraints)
                if objective(x, e, T, e_dot, sigma_test) >= objective(xb, e, T, e_dot, sigma_test):
                    break
            x = deepcopy(xb)
        else:
            s = alpha * s
    return xb


def trial(x, s, constraints):
    for j in range(len(x)):
        trial_x = deepcopy(x)
        trial_x[j] = trial_x[j] + s * constraints[1][j]
        if constraints[1][j] > trial_x[j] > constraints[0][j] and \
                objective(trial_x, e, T, e_dot, sigma_test) < objective(x, e, T, e_dot, sigma_test):
            x = deepcopy(trial_x)
        else:
            trial_x[j] = trial_x[j] - 2 * s * constraints[1][j]
            if constraints[1][j] > trial_x[j] > constraints[0][j] and \
                    objective(trial_x, e, T, e_dot, sigma_test) < objective(x, e, T, e_dot, sigma_test):
                x = deepcopy(trial_x)
    # print(x)
    return x


# load Excel file and initialize arrays
# wb = load_workbook('./Doswiadczenia.xlsx', read_only=False)
# e = []
# sigma_test = []
# T = []
# e_dot = []

# set experimental values
# for ws in wb.worksheets:
#     e_values = [ws.cell(i, 1).value for i in range(2, 23)]
#     sigma_values = [ws.cell(i, 2).value for i in range(2, 23)]
#     T.append(ws.cell(2, 4).value)
#     e_dot.append(ws.cell(2, 5).value)
#     e.append(e_values)
#     sigma_test.append(sigma_values)

