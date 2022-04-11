from openpyxl import load_workbook
import random

# This code takes Excel spreadsheet as an input and outputs data used in simulations
# To use different inputs, comment or uncomment as needed

# SIM 2
wb = load_workbook('./Doswiadczenia-v03.xlsx')
sigma_test = []
e = []
T_def = []
e_dot = []
q_def = 312500
R_gas = 8.314

for ws in wb.worksheets:
    e_values = [ws.cell(i, 1).value for i in range(4, 55)]
    sigma_values = [ws.cell(i, 2).value for i in range(4, 55)]
    T_def.append(ws.cell(2, 1).value)
    e_dot.append(ws.cell(2, 2).value)
    e.append(e_values)
    sigma_test.append(sigma_values)

num_of_experiments = len(T_def)
num_of_measurements = len(e[0])

smallest = 10000000000000.0
constraints = [[1E12, 3.0, 1E-04, 1E11, 3.0, 1E-04, 0.01, 1E-06, 1E12, 2.0, 1E-04, 1E-06, 0.0, 1.0, -1.0],
               [1E14, 6.0, 0.1, 1E13, 6.0, 0.01, 0.5, 0.001, 1E14, 5.0, 0.01, 0.1, 1.0, 50.0, 0.0]]
a = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
for i in range(15):
    a[i] = random.uniform(constraints[0][i], constraints[1][i])

s = 0.02
# 0 < alpha < 1
alpha = 0.5
# precision
epsilon = 0.000001
